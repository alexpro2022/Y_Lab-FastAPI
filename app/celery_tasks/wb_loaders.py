from gspread_asyncio import AsyncioGspreadClientManager, AsyncioGspreadSpreadsheet
from openpyxl import load_workbook
from google.oauth2.service_account import Credentials
from googleapiclient import discovery

from app.core.config import settings
from pathlib import Path
from datetime import datetime as dt
import httpx
import logging
logging.basicConfig(level=logging.INFO)

FILE_PATH = Path('admin/Menu.xlsx')

service_account_info = {
    'type': settings.type,
    'project_id': settings.project_id,
    'private_key_id': settings.private_key_id,
    'private_key': settings.private_key,
    'client_email': settings.client_email,
    'client_id': settings.client_id,
    'auth_uri': settings.auth_uri,
    'token_uri': settings.token_uri,
    'auth_provider_x509_cert_url': settings.auth_provider_x509_cert_url,
    'client_x509_cert_url': settings.client_x509_cert_url,
}


def get_creds() -> Credentials:
    creds = Credentials.from_service_account_info(service_account_info)
    Credentials.token_state
    scoped = creds.with_scopes([
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ])
    return scoped


agcm = AsyncioGspreadClientManager(get_creds)


async def get_ss() -> AsyncioGspreadSpreadsheet:
    agc = await agcm.authorize()
    return await agc.open(settings.spreadsheet_title)


async def google_load_workbook() -> list[list[str]]:
    ss = await get_ss()
    ws = await ss.get_sheet1()
    return await ws.get_all_values()


def local_load_workbook():
    return load_workbook(FILE_PATH)['Лист1'].values


async def google_get_mod_timestamp() -> float:
    '''ss = await get_ss()
    revisions_uri = f'https://www.googleapis.com/drive/v3/files/{ss.id }/revisions'
    headers = {'Authorization': f'Bearer {creds.token}'}
    async with httpx.AsyncClient as client:
        response = await client.get(revisions_uri, headers=headers).json()
    logging.info(response)  # ['revisions'][-1]['modifiedTime'])
    raise ValueError(response) '''
    return dt.now().timestamp()


async def is_modified() -> bool:
    if settings.google_sheets:
        timestamp = await google_get_mod_timestamp()
    else:
        timestamp = FILE_PATH.stat().st_mtime
    return (dt.now() - dt.fromtimestamp(timestamp)).total_seconds() <= settings.celery_task_period
